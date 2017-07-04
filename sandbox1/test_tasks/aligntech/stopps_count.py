__author__ = 'atatarnikov'

from openpyxl import load_workbook
from geopy.distance import vincenty
import time
from decimal import Decimal

data_path = './data/'

subway = load_workbook(data_path + 'subway.xlsx', read_only=True)
surface = load_workbook(data_path + 'surface.xlsx', read_only=True)

subway_ws = subway.active
surface_ws = surface.active

# !! it is a pity - the script shows poor performance
# no glue for the moment to speed it up ))=^

# the algorithm considers each subway station exit as the station,
# otherwise, additional post-processing required

# subway_ws['B2'].value  name
# subway_ws['C2'].value  dolgota
# subway_ws['D2'].value  shirota

# surface_ws['B2'].value  name
# surface_ws['C2'].value  dolgota
# surface_ws['D2'].value  shirota


def distance(d1, s1, d2, s2):
    return vincenty((d1, s1), (d2, s2)).kilometers


freq_subway = []

subway_idx = 2

start = time.time()

# form a list of surface coordinates

surf_coord = []
surf_idx1 = 2

surf_time1 = time.time()

while True:
    str_surf_idx1 = str(surf_idx1)

    # TODO: remove second condition when get completed
    if surface_ws['C' + str_surf_idx1].value is None or surf_idx1 > 200:
        break

    surf_coord.append((Decimal(surface_ws['C' + str_surf_idx1].value),
                       Decimal(surface_ws['D' + str_surf_idx1].value)))
    surf_idx1 += 1

surf_time2 = time.time()

print 'Forming surf coords list taken: ' + str(surf_time2 - surf_time1)

coords_delta = Decimal(0.01)

# the loop over all subway stations
while True:
    surface_idx = 2
    curr_station_stop_count = 0

    sub_idx = str(subway_idx)

    if subway_ws['C' + sub_idx].value is None:
        break

    print 'Current subway idx: ' + sub_idx

    subway_curr_coord_pair = (Decimal(subway_ws['C' + sub_idx].value),
                              Decimal(subway_ws['D' + sub_idx].value))

    # filter surf_coord list in dependence of current subway station coords

    narrow_surf_coords = [x for x in surf_coord
                          if abs(x[0] - subway_curr_coord_pair[0]) < coords_delta
                          or abs(x[1] - subway_curr_coord_pair[1]) < coords_delta]

    for x in narrow_surf_coords:
        if distance(subway_curr_coord_pair[0], subway_curr_coord_pair[1],
                    x[0], x[1]) <= 0.5:
            curr_station_stop_count += 1

    # while True:
    # sur_idx = str(surface_idx)
    #
    #     # print 'Current surface idx: ' + sur_idx
    #
    #     if surface_ws['C' + sur_idx].value is None:
    #         break
    #
    #     if distance(subway_ws['C' + sub_idx].value, subway_ws['D' + sub_idx].value,
    #                 surface_ws['C' + sur_idx].value, surface_ws['D' + sur_idx].value) <= 0.5:
    #         curr_station_stop_count += 1
    #
    #     surface_idx += 1

    freq_subway.append((curr_station_stop_count, subway_ws['B' + sub_idx].value))
    subway_idx += 1

try:
    # was not able to find pytonic way to make such output )=^
    sorted_freq_subway = sorted(freq_subway, reverse=True)

    max_val = sorted_freq_subway[0][0]  # get max value

    # try to print equal-to-max elements
    for elm in sorted_freq_subway:
        if elm[0] != max_val:
            break

        print elm

except IndexError:
    print 'Freq array is empty'

end = time.time()
print 'Time taken: ' + str(end - start)